"""file_readers — read files into (content, channel-type) for the judge.

xlsx reads preserve cell style (font/size/color/fill/bold/format/merge/
cached formula value); see xlsx_utils.py + read_xlsx below.
extract_model_response concatenates assistant content, splits on </think> and
</longcat_think> taking the last segment, and stops at "Read HEARTBEAT.md".

file2type values feed build_prompt's channel routing:
  "text" | "ori_path" | "trans_path/compress" | "trans_path/pdf".
"""

import os
import csv
import io
from PIL import Image

from .xlsx_utils import (
    _get_xlsx_theme_fonts, _get_xlsx_theme_colors,
    _xlsx_cell_to_text, _display_width, _pad,
)
from .pdf_utils import (
    _check_pdf_valid, _check_page_limit, _convert_to_pdf_via_libreoffice,
)
from .img_utils import _compress_image


def list_all_files(dir_path):
    """递归遍历文件夹，返回其下所有文件的绝对路径列表。"""
    abs_dir = os.path.abspath(dir_path)
    file_list = []
    for root, dirs, files in os.walk(abs_dir):
        SKIP_DIRS = {'__pycache__', 'node_modules', '.git', '__MACOSX', '.trans'}
        dirs[:] = [d for d in dirs if not d.startswith('.') and d not in SKIP_DIRS]
        for filename in files:
            if filename.startswith('.') or filename.endswith(('.pyc', '.pyo')):
                continue
            file_list.append(os.path.join(root, filename))

    print(f"{dir_path}目录下已找到 {len(file_list)} 个文件：")
    for f in file_list:
        print(f" - {f}")
    return file_list


def extract_files(file_paths, use_pdf_cache=True):
    """根据文件路径列表读取每个文件，返回 (file2content, file2type)。

    路径以 "*" 结尾表示递归读取该目录下所有文件。任一路径不存在则返回 None。
    """
    file2content = {}
    file2type = {}
    for file_path in file_paths:
        if file_path.endswith("*"):
            child_files = list_all_files(file_path[:-1])
            for child_file in child_files:
                assert os.path.exists(child_file)
                file2content[child_file], file2type[child_file] = read_file(child_file, use_pdf_cache=use_pdf_cache)
        else:
            try:
                assert os.path.exists(file_path)
                file2content[file_path], file2type[file_path] = read_file(file_path, use_pdf_cache=use_pdf_cache)
            except AssertionError:
                print(f"警告: 文件路径 {file_path} 不存在。")
                return None
    return file2content, file2type


def extract_model_response(agent_messages):
    """从 agent_messages 提取模型最终 response 文本。"""
    response_parts = []
    for msg in agent_messages:
        if msg.get("role") == "assistant" and "content" in msg:
            assert isinstance(msg["content"], str) or msg["content"] is None
            if msg["content"] is not None:
                content = msg["content"]
                content = content.split("</think>")[-1]
                content = content.split("</longcat_think>")[-1]
                content = content.strip()
                if content:
                    response_parts.append(content)
        if msg.get("role") == "user" and "content" in msg:
            user_content = msg["content"]
            if isinstance(user_content, str) and "Read HEARTBEAT.md" in user_content:
                break
    return "\n".join(response_parts)


# ============================================================
# 文件读取
# ============================================================

def read_file(file_path: str, use_pdf_cache: bool = True):
    """根据扩展名分派读取函数，返回 (content, file2type)。"""
    _, ext = os.path.splitext(file_path)
    ext = ext.lower()

    lo_formats = {".docx", ".doc", ".pptx"}

    reader_map = {
        ".xlsx": read_xlsx,
        ".csv":  read_csv,
        ".pdf":  read_pdf,
        ".docx": read_docx,
        ".doc":  read_doc,
        ".pptx": read_pptx,
        ".png":  read_image,
        ".jpg":  read_image,
        ".jpeg": read_image,
        ".svg":  read_svg,
    }

    reader = reader_map.get(ext)
    if reader is not None:
        if ext in lo_formats:
            read_content = reader(file_path, use_pdf_cache=use_pdf_cache)
        else:
            read_content = reader(file_path)
    else:
        read_content = read_plain_text(file_path)

    if ext in [".png", ".jpg", ".jpeg", ".pdf"]:
        if ext in [".png", ".jpg", ".jpeg"] and os.path.abspath(read_content) != os.path.abspath(file_path):
            assert "compress" in read_content, f"图片 {file_path} 的读取结果 {read_content} 不包含 'compress'，不符合预期的压缩路径格式。"
            return read_content, "trans_path/compress"
        return read_content, "ori_path"
    if ext in [".doc", ".docx", ".pptx"]:
        return read_content, "trans_path/pdf"
    return read_content, "text"


# -------------------- 纯文本 --------------------

def read_plain_text(file_path: str) -> str:
    """以纯文本方式读取文件，依次尝试 UTF-8 → GBK，均失败则用 UTF-8 lossy 兜底。"""
    for enc in ("utf-8", "gbk"):
        try:
            with open(file_path, "r", encoding=enc) as f:
                return f.read()
        except (UnicodeDecodeError, UnicodeError):
            continue
    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
        return f.read()


# -------------------- 需要专用库的格式 --------------------

def read_xlsx(file_path: str) -> str:
    """读取 Excel (.xlsx)，返回带表格格式和单元格样式信息的纯文本。"""
    from openpyxl import load_workbook

    wb = load_workbook(file_path, data_only=False)
    wb_data = load_workbook(file_path, data_only=True)

    theme_fonts = _get_xlsx_theme_fonts(file_path)
    theme_colors = _get_xlsx_theme_colors(file_path)
    default_font_name = None
    if wb._fonts and len(wb._fonts) > 0:
        default_font_name = wb._fonts[0].name

    sheet_texts = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws_data = wb_data[sheet_name] if sheet_name in wb_data.sheetnames else None
        if ws.max_row is None or ws.max_column is None:
            continue

        merged_map = {}
        for merged_range in ws.merged_cells.ranges:
            cells = list(merged_range.cells)
            if cells:
                origin = cells[0]
                merged_map[origin] = "merged_origin"
                for rc in cells[1:]:
                    merged_map[rc] = "merged_hidden"

        rows = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            row_data = []
            for cell in row:
                rc = (cell.row, cell.column)
                if merged_map.get(rc) == "merged_hidden":
                    row_data.append("")
                    continue
                cached_val = ws_data.cell(row=cell.row, column=cell.column).value if ws_data else None
                row_data.append(_xlsx_cell_to_text(
                    cell, theme_fonts=theme_fonts,
                    default_font_name=default_font_name,
                    theme_colors=theme_colors,
                    cached_value=cached_val))
            rows.append(row_data)

        if not rows:
            continue

        num_cols = max(len(r) for r in rows)
        for r in rows:
            while len(r) < num_cols:
                r.append("")
            for i in range(len(r)):
                r[i] = r[i].replace("|", "\\|")

        col_widths = []
        for col_idx in range(num_cols):
            max_w = 0
            for r in rows:
                w = _display_width(r[col_idx])
                if w > max_w:
                    max_w = w
            col_widths.append(max(max_w, 2))

        lines = []
        if len(wb.sheetnames) > 1:
            lines.append(f"## Sheet: {sheet_name}")
            lines.append("")

        for row_idx, row in enumerate(rows):
            cells = [_pad(row[col], col_widths[col]) for col in range(num_cols)]
            line = "| " + " | ".join(cells) + " |"
            lines.append(line)
            if row_idx == 0:
                sep = "|" + "|".join("-" * (w + 2) for w in col_widths) + "|"
                lines.append(sep)

        sheet_texts.append("\n".join(lines))

    wb.close()
    wb_data.close()
    return "\n\n".join(sheet_texts)


def read_csv(file_path: str) -> str:
    """读取 CSV 文件，返回带表格格式的纯文本。"""
    raw_text = read_plain_text(file_path)
    if raw_text.startswith('﻿'):
        raw_text = raw_text[1:]
    raw_text = raw_text.replace('\x00', '')

    try:
        dialect = csv.Sniffer().sniff(raw_text[:8192])
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ","

    reader = csv.reader(io.StringIO(raw_text), delimiter=delimiter)
    rows = [row for row in reader if any(cell.strip() for cell in row)]

    if not rows:
        return ""

    num_cols = max(len(row) for row in rows)
    for row in rows:
        while len(row) < num_cols:
            row.append("")
        for i in range(len(row)):
            row[i] = row[i].strip().replace("|", "\\|")

    col_widths = []
    for col_idx in range(num_cols):
        max_w = 0
        for row in rows:
            w = _display_width(row[col_idx])
            if w > max_w:
                max_w = w
        col_widths.append(max(max_w, 2))

    lines = []
    for row_idx, row in enumerate(rows):
        cells = [_pad(row[col], col_widths[col]) for col in range(num_cols)]
        line = "| " + " | ".join(cells) + " |"
        lines.append(line)
        if row_idx == 0:
            sep = "|" + "|".join("-" * (w + 2) for w in col_widths) + "|"
            lines.append(sep)

    return "\n".join(lines)


def read_pdf(file_path: str) -> str:
    """读取 PDF 文件，校验后返回文件的绝对路径。"""
    _check_pdf_valid(file_path)
    return os.path.abspath(file_path)


def read_docx(file_path: str, use_pdf_cache: bool = True) -> str:
    """将 Word (.docx) 无损转换为 PDF，返回 PDF 的绝对路径。"""
    _check_page_limit(file_path)
    return _convert_to_pdf_via_libreoffice(file_path, use_pdf_cache=use_pdf_cache)


def read_doc(file_path: str, use_pdf_cache: bool = True) -> str:
    """将旧版 Word (.doc) 无损转换为 PDF，返回 PDF 的绝对路径。"""
    _check_page_limit(file_path)
    return _convert_to_pdf_via_libreoffice(file_path, use_pdf_cache=use_pdf_cache)


def read_pptx(file_path: str, use_pdf_cache: bool = True) -> str:
    """将 PowerPoint (.pptx) 无损转换为 PDF，返回 PDF 的绝对路径。"""
    _check_page_limit(file_path)
    return _convert_to_pdf_via_libreoffice(file_path, use_pdf_cache=use_pdf_cache)


def read_image(file_path: str) -> str:
    """读取图片 (.png/.jpg)，校验完整性后返回路径（超 5MB 则返回压缩路径）。"""
    abs_path = os.path.abspath(file_path)
    if not os.path.isfile(abs_path):
        raise FileNotFoundError(f"图片文件不存在: {abs_path}")

    try:
        with Image.open(abs_path) as img:
            img.verify()
    except Exception as e:
        raise ValueError(f"图片文件损坏 (verify 失败): {abs_path}\n原始错误: {e}")

    try:
        with Image.open(abs_path) as img:
            img = img.convert("RGB")
            img.load()
    except Exception as e:
        raise ValueError(f"图片文件损坏 (像素解码失败): {abs_path}\n原始错误: {e}")

    abs_path = _compress_image(abs_path)
    return abs_path


def read_svg(file_path: str) -> str:
    """以纯文本方式读取 SVG 文件（SVG 本质是 XML 文本）。"""
    return read_plain_text(file_path)
